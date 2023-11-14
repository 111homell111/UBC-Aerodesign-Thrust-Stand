from openpyxl import Workbook
from openpyxl import load_workbook
import os
from datetime import datetime

first = 0
__location__ = os.path.realpath(

    os.path.join(os.getcwd(), os.path.dirname(__file__)))




import random
import tkinter as Tk
from itertools import count

import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.pyplot import figure

plt.style.use('fivethirtyeight')
# values for first graph
x_vals = []
y_vals = [0,0,5,0,0]
# values for second graph
y_vals2 = [0,0,5,0,0]

index = count()
index2 = count()

def animate(i):
    # Generate values
    x_vals.append(next(index))
    y_vals.append(random.randint(0, 5))
    y_vals2.append(random.randint(0, 5))
    y_vals = y_vals[1:]
    y_vals2 = y_vals2[1:]
    # Get all axes of figure
    ax1, ax2 = plt.gcf().get_axes()
    # Clear current data
    ax1.cla()
    ax2.cla()
    # Plot new data
    ax1.plot.set(x_vals, y_vals)
    ax2.plot.set(x_vals, y_vals2)
    ax1.set_xlim(x_vals[0], x_vals[-1])
    ax1.set_xlim(x_vals[0], x_vals[-1])
    canvas.draw_idle()  # redraw plot

# GUI
root = Tk.Tk()
label = Tk.Label(root, text="Realtime Animated Graphs").grid(column=0, row=0)
figure(figsize=(8, 6), dpi=80)

# graph 1
canvas = FigureCanvasTkAgg(plt.gcf(), master=root, )
canvas.get_tk_widget().grid(column=0, row=1)
# Create two subplots in row 1 and column 1, 2
plt.gcf().subplots(1, 2)
ani = FuncAnimation(plt.gcf(), animate, interval=1000, blit=False)


Tk.mainloop()

while True():
    x = input("GIVE: ")
    print(x)


'''
wb = load_workbook(filename = (os.path.join(__location__, 'Test_data')+'.xlsx'))
ws = wb['Manual']
#current_row = ws.max_row + 1
 
#thrusts = []

for i in range (9): #Thrust
    
    if first == 0:
        ws.append([datetime.today().strftime('%Y-%m-%d %H:%M:%S')])
        first = 1

    throttle = i
    thrust = i+1
    current = 999999

    ws.append([throttle, thrust, current])  

    wb.save(filename=os.path.join(__location__, 'Test_data')+'.xlsx')
    wb.close()
'''