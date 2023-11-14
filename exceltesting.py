from openpyxl import Workbook
from openpyxl import load_workbook
import os

__location__ = os.path.realpath(
    os.path.join(os.getcwd(), os.path.dirname(__file__)))

wb = load_workbook(filename = (os.path.join(__location__, 'Test_data')+'.xlsx'))
ws = wb.active
ws['A1'] = 42
wb.save((os.path.join(__location__, 'Test_data'+'.xlsx')))
