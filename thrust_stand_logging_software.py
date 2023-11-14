'''
    Purpose: I hate APSC 160
    Purpose: Control thrust stand dfj;;lkafslkjfjdkla;fadf;jdsfj;ksfjk;la

'''



'''-----------------Import Libraries----------------'''
from curses import window
import tkinter as tk
from tkinter import ttk
from tkinter import simpledialog
import serial
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import asyncio
from datetime import datetime
from tkinter import filedialog as fd

#Will need these bottom two imports to plot/display realtime data
#import matplotlib.pyplot as plt
#import matplotlib.animation as animation


'''---------------Set Up Arduino -----------------'''
#port = '/dev/cu.usbmodem11301' #You can find the port on the bottom right of
port = '/dev/cu.usbserial-1130' #You can find the port on the bottom right of
baud_rate = 9600 #Idk it just has to match the Arduino's code (Prolly use 115200 or 9600 for now)
#arduino = serial.Serial(port, baud_rate)

    



'''-------------------??? :D---------------------'''
__location__ = os.path.realpath(
    os.path.join(os.getcwd(), os.path.dirname(__file__)))
WINDOW_WIDTH = "1000"
WINDOW_HEIGHT = "800"
BUTTON_HEIGHT = 3
BUTTON_WIDTH = 10
MODES = ["Automatic", "Manual", "Timed"]
NUMREADINGS = 5
first = True


'''------The Real Stuff--------'''
class Thrust_Gui:
    def __init__(self, window):
        self.mode = MODES[0]


        '''-------------Create Main Window-------------'''
        self.window = window 
        self.title = window.title("Thrust Stand Controller")
        self.size = window.geometry(WINDOW_WIDTH + "x" + WINDOW_HEIGHT)


        '''-------------Create Frames-------------'''
        self.left_frame = tk.Frame(window, width = 400, height = 785, bg = "pink")
        self.left_frame.grid(row=0, column=0, padx=10, pady=10)


        '''------------------Create Widgets-----------------'''
        self.infile_button = tk.Button(self.left_frame, command = self.select_infile, text = "Select input file")
        self.infile_display = tk.Text(self.left_frame, width = 15, height = 1.495)
        self.outfiel_button = tk.Button(self.left_frame, command = self.select_outfile, text = "Select output file")

        self.mode_value = tk.StringVar(window)
        self.mode_value.set(self.mode)
        self.select_mode = tk.OptionMenu(window, self.mode_value, *MODES, command = self.change_mode)

        self.start_button = tk.Button(window, command = self.start, text = "Start" )
        self.readings_button = tk.Button(window, command = self.get_readings, text = "Get Readings" )
        self.motor_slider = tk.Scale(window, from_ = 0, to = 1023, orient = 'vertical', command = self.slider, length = 200, label = "Motor")

        self.window.bind("<Up>", lambda e: self.motor_slider.set(self.motor_slider.get()-10))
        self.window.bind("<Down>", lambda e: self.motor_slider.set(self.motor_slider.get()+10))
        
        self.text=tk.Text(window, width=60, height=15, state = "disabled")
        
        '''------------------Format Widgets-------------------'''
        self.start_button.config(width = BUTTON_WIDTH, height = BUTTON_HEIGHT)


        '''----------------Place Widgets-----------------'''
        self.infile_button.place(relx = 0.05, rely = 0.3)
        self.infile_display.place(relx = 0.45, rely = 0.305)

        self.select_mode.place(relx = 0.05, rely = 0.05)
        self.motor_slider.place(relx = 0.45, rely = 0.3)
        self.start_button.place(relx = 0.05, rely = 0.15)
        self.readings_button.place(relx = 0.55, rely= 0.55)
        self.text.place(relx = 0.55, rely = 0.3)
       


        '''-----------------Work Book------------------'''
        #file path idfk

        #wb = load_workbook(filename = (os.path.join(__location__, 'Test_data')+'.xlsx'))
        #ws = wb.active
        #ws['A1'] = 42
        #wb.save((os.path.join(__location__, 'Test_data'+'.xlsx')))


        

    '''-------------------Create Functions------------------'''
    def start(self):
        '''
        Run a series of start checks, if passed, run test
        '''

        if self.mode == 'Automatic':
            if self.file_menu.get() == '':
                tk.messagebox.showinfo (title = 'Error', message = 'No input file selected!', icon = 'warning')
                return 0
            input_file = open(os.path.join(__location__, self.file_menu.get()), 'r') #read input file
            input_list = [line .rstrip('\n') for line in input_file.readlines()]
            input_list.append("000000")
        #Would prefer to ask for a file name first
        wb = load_workbook(filename = (os.path.join(__location__, 'Test_data')+'.xlsx'))
        ws = wb.active
        for i in range (len(input_list)):
            line = input_list[i]
            duration = int(line[3:])
            throttle = int(line[:3])
            self.window.after(duration * 1000, self.send_throttle(throttle))
            print('Test: ', line)   
            data = arduino.readline()
            print(data)
            ws['A' + str(i+1)] = line
            ws['B' + str(i+1)] = data
            #ws['C' + str(i+1)] = data[1]
        print('finished')
        wb.save((os.path.join(__location__, 'Test_data'+'.xlsx')))




    def send_throttle(self, throttle):
        string = 'X{0:d}'.format(throttle)
        arduino.write(string.encode('utf-8'))

    def stop(self):
        string = 'X0'
        arduino.write(string.encode('utf-8'))
        
    def select_infile(self):
        filename = fd.askopenfilename(filetypes=[("text file","*.txt")])
        lastslash = filename.rindex("/") + 1
        self.infile_display.insert(tk.END, filename[lastslash:])
    
    def get_readings(self):
        time.sleep(1);
        string = 'S'

        arduino.write(string.encode('utf-8'))
        wb = load_workbook(filename = (os.path.join(__location__, 'Test_data')+'.xlsx'))
        ws = wb['Manual']

        if first:
            ws.append([datetime.today().strftime('%Y-%m-%d %H:%M:%S')])
            first = False 

        for i in range (NUMREADINGS): #Thrust
            arduino.readline() #read title
            throttle = (arduino.readline()).decode('utf-8')
            throttle = throttle.strip() 

            thrust = (arduino.readline()).decode('utf-8')
            thrust = thrust.strip()

            current = (arduino.readline()).decode('utf-8')
            current = thrust.strip()

            self.text.insert(tk.END, str(i+1) + ": " + thrust + "\n")
            ws.append([throttle, thrust, current])
        
        wb.save(filename=os.path.join(__location__, 'Test_data')+'.xlsx')
        wb.close()


    def change_mode(self, new_mode):
        if new_mode == MODES[0]:
            self.file_menu["state"] = "readonly"
        elif new_mode == MODES[1]:
            self.file_menu["state"] = "disabled"
        elif new_mode == MODES[2]:
            self.file_menu["state"] = "disabled"
        self.mode = new_mode
        
    def slider(self, throttle):
        throttle = int(throttle) #for slider LOL or like just in case ig :) i would prefer to just pass ints... should i>??????a number
        string = 'X{0:d}'.format(throttle)
        arduino.write(string.encode('utf-8'))
        print(string)

    def save(self):
        pass
        #write to file lmfaooooo


'''------------------Run Code---------------'''
#def exit(): #This is very much a useless roundabout way but like whatever works works
#   string = 'X0'
#    arduino.write(string.encode('utf-8'))
#    root.destroy()
root = tk.Tk()
app = Thrust_Gui(root)
#root.protocol("WM_DELETE_WINDOW", exit)
root.mainloop()



'''
NOTES:
- Uh calibrate by doing 100% then 0% and wait for the beeps lolol
- The lag is caused by keyboard inputs :,)





left and right arrow key to control sldier

'''


'''
create a pop up window
    def save(self):
        popup = tk.Tk()
        popup.wm_title("!")
        label = ttk.Label(popup, text=msg, font=NORM_FONT)
        label.pack(side="top", fill="x", pady=10)
        B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
        B1.pack()
        popup.mainloop()
        '''

'''
          if self.output_menu.get() == '':
                self.no_output_message = tk.messagebox.askquestion(title = 'Warning', message = 'No output file selected. Proceed anyway? (Will create a new file)', icon = 'info')
                if self.no_output_message == 'No':
                    return 0
                elif self.no_output_message == 'Yes':
                    output_name = simpledialog.askstring(title = 'File Name', message = 'Please name your file (omit.xlsx)')
                    #IDK?????? LMAOOOO
                    
'''


                #Unrelated by when do I use a switch statement over elif

